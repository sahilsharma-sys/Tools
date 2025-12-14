# ==========================================================
# 🧰 UNIFIED HELPER + Master PORTAL
# Author: Sahil (Unified Version)
# ==========================================================

import streamlit as st
import pandas as pd
import os, io, zipfile, shutil, re, requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from math import radians, sin, cos, sqrt, atan2
from concurrent.futures import ThreadPoolExecutor

# ==========================================================
# PAGE CONFIG (ONLY ONCE)
# ==========================================================
st.set_page_config(
    page_title="🧰 Helper Portal",
    layout="wide"
)

st.title("🧰 Helper Portal")
st.caption("Single App • All  • No Duplication")

# ==========================================================
# SIDEBAR
# ==========================================================
st.sidebar.title("📌 Master Point")

section = st.sidebar.radio(
    "Select Section",
    ["🏠 Home", "📦 Sahil Helper Portal", "🛠 Master"]
)

tool = None

if section == "📦 Sahil Helper Portal":
    tool = st.sidebar.selectbox(
        "Choose Tool",
        [
            "Data Compiler",
            "Files Splitter",
            "Pincode Zone + Distance",
            "Data Cleaner & Summary",
            "Create Folders from List"
        ]
    )

if section == "🛠 Master":
    tool = st.sidebar.selectbox(
        "Choose Tool",
        [
            "New File Creation",
            "CSV → XLSX Converter",
            "Merchant Auto Rename",
            "Delete Files Tool",
            "Excel Formula Updater",
            "Courier Cost Updater"
        ]
    )

# ==========================================================
# COMMON UTILITIES
# ==========================================================
def list_files(path):
    if not os.path.exists(path):
        return []
    return os.listdir(path)

# ---------- PINCODE UTILITIES ----------
METRO_RANGES = [
    range(110001,110099), range(400001,400105), range(700001,700105),
    range(600001,600119), range(560001,560108), range(500001,500099),
    range(380001,380062), range(411001,411063), range(122001,122019)
]

def is_metro(pin):
    try:
        return any(int(pin) in r for r in METRO_RANGES)
    except:
        return False

def get_location(pin):
    try:
        r = requests.get(f"https://api.postalpincode.in/pincode/{pin}", timeout=10)
        d = r.json()
        if d[0]["Status"] == "Success":
            po = d[0]["PostOffice"][0]
            return po["District"], po["State"]
    except:
        pass
    return "N/A", "N/A"

def get_latlon(pin):
    try:
        url = f"https://nominatim.openstreetmap.org/search?postalcode={pin}&country=India&format=json"
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}).json()
        if r:
            return float(r[0]["lat"]), float(r[0]["lon"])
    except:
        pass
    return None, None

def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    return round(R * 2 * atan2(sqrt(a), sqrt(1-a)), 2)

# ==========================================================
# HOME
# ==========================================================
if section == "🏠 Home":
    c1, c2, c3 = st.columns(3)
    c1.metric("📦 Helper ", 5)
    c2.metric("🛠 Excel ", 6)
    c3.metric("🔥 Total ", 11)


# ==========================================================
# ===================== Sahil Helper Portal ======================
# ==========================================================
if section == "📦 Sahil Helper Portal":

    # ---------- DATA COMPILER ----------
    if tool == "Data Compiler":
        st.header("📁 Data Compiler")
        files = st.file_uploader(
            "Upload CSV / Excel Files",
            type=["csv","xlsx"],
            accept_multiple_files=True
        )

        if files:
            dfs = []
            for f in files:
                df = pd.read_csv(f) if f.name.endswith(".csv") else pd.read_excel(f)
                df["Source File"] = f.name
                dfs.append(df)

            final_df = pd.concat(dfs, ignore_index=True)
            st.dataframe(final_df, use_container_width=True)

            st.download_button(
                "⬇ Download Compiled CSV",
                final_df.to_csv(index=False).encode(),
                "compiled.csv"
            )

    # ---------- FILES SPLITTER ----------
    elif tool == "Files Splitter":
        st.header("📂 Files Splitter")
        f = st.file_uploader("Upload CSV / XLSX", type=["csv","xlsx"])

        if f:
            df = pd.read_csv(f) if f.name.endswith(".csv") else pd.read_excel(f)
            col = st.selectbox("Split by Column", df.columns)

            if st.button("🚀 Split & Download"):
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as z:
                    for v in df[col].dropna().unique():
                        z.writestr(
                            f"{v}.csv",
                            df[df[col] == v].to_csv(index=False)
                        )

                st.download_button(
                    "⬇ Download ZIP",
                    zip_buffer.getvalue(),
                    "split_files.zip"
                )

    # ---------- PINCODE ZONE + DISTANCE ----------
    elif tool == "Pincode Zone + Distance":
        st.header("📍 Pincode Zone + Distance")
        txt = st.text_area("Enter From,To Pincode (one per line)", height=200)

        if st.button("🚀 Calculate"):
            rows = []
            for line in txt.splitlines():
                if "," in line:
                    fpin, tpin = line.split(",")
                    fdist, fstate = get_location(fpin.strip())
                    tdist, tstate = get_location(tpin.strip())

                    lat1, lon1 = get_latlon(fpin.strip())
                    lat2, lon2 = get_latlon(tpin.strip())
                    km = haversine(lat1, lon1, lat2, lon2) if None not in [lat1,lon1,lat2,lon2] else "N/A"

                    if fpin == tpin:
                        zone = "LOCAL"
                    elif is_metro(fpin) and is_metro(tpin):
                        zone = "METRO"
                    elif fstate == tstate:
                        zone = "REGIONAL"
                    else:
                        zone = "ROI"

                    rows.append([fpin, tpin, fdist, fstate, tdist, tstate, km, zone])

            df = pd.DataFrame(
                rows,
                columns=["From Pincode","To Pincode","From District","From State","To District","To State","Distance KM","Zone"]
            )

            st.dataframe(df, use_container_width=True)
            st.download_button(
                "⬇ Download CSV",
                df.to_csv(index=False).encode(),
                "pincode_zone_distance.csv"
            )

    # ---------- DATA CLEANER ----------
    elif tool == "Data Cleaner & Summary":
        st.header("🧹 Data Cleaner & Summary")
        f = st.file_uploader("Upload CSV / XLSX", type=["csv","xlsx"])

        if f:
            df = pd.read_csv(f) if f.name.endswith(".csv") else pd.read_excel(f)
            cleaned = df.drop_duplicates().dropna(how="all")

            st.metric("Original Rows", len(df))
            st.metric("Cleaned Rows", len(cleaned))

            st.dataframe(cleaned, use_container_width=True)
            st.download_button(
                "⬇ Download Cleaned File",
                cleaned.to_csv(index=False).encode(),
                "cleaned.csv"
            )

    # ---------- CREATE FOLDERS ----------
    elif tool == "Create Folders from List":
        st.header("📂 Create Folders from List")
        names = st.text_area("Enter folder names (one per line)")

        if st.button("🚀 Generate ZIP"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as z:
                for n in names.splitlines():
                    z.writestr(f"{n}/", "")
            st.download_button(
                "⬇ Download ZIP",
                zip_buffer.getvalue(),
                "folders.zip"
            )

# ==========================================================
# ================= Master ===================
# ==========================================================
if section == "🛠 Master":

    # ---------- NEW FILE CREATION ----------
    if tool == "New File Creation":
        st.header("📂 New File Creation")
        src = st.text_input("Source Folder Path")
        dest = st.text_input("Destination Folder Path")

        if st.button("🚀 Create New Files"):
            os.makedirs(dest, exist_ok=True)
            for f in list_files(src):
                shutil.copy(os.path.join(src,f), os.path.join(dest,f))
            st.success("✅ Files copied successfully")

    # ---------- CSV TO XLSX ----------
    elif tool == "CSV → XLSX Converter":
        st.header("🔄 CSV → XLSX Converter")
        path = st.text_input("Folder Path")

        if st.button("🚀 Convert"):
            for f in list_files(path):
                if f.lower().endswith(".csv"):
                    pd.read_csv(os.path.join(path,f)).to_excel(
                        os.path.join(path, f.replace(".csv",".xlsx")),
                        index=False
                    )
            st.success("✅ Conversion completed")

    # ---------- MERCHANT AUTO RENAME ----------
    elif tool == "Merchant Auto Rename":
        st.header("📝 Merchant Auto Rename")
        path = st.text_input("Folder Path")

        if st.button("🚀 Rename"):
            renamed = []
            for f in list_files(path):
                if f.endswith(".xlsx"):
                    wb = load_workbook(os.path.join(path,f))
                    ws = wb.active
                    if str(ws["A1"].value).lower() == "client name":
                        new_name = str(ws["A2"].value) + ".xlsx"
                        wb.close()
                        os.rename(
                            os.path.join(path,f),
                            os.path.join(path,new_name)
                        )
                        renamed.append(new_name)

            st.dataframe(pd.DataFrame(renamed, columns=["Renamed Files"]))

    # ---------- DELETE FILES ----------
    elif tool == "Delete Files Tool":
        st.header("🗑 Delete Files Tool")
        folder = st.text_input("Target Folder Path")
        file = st.file_uploader("Upload Excel with file names")

        if file and st.button("🚀 Delete"):
            df = pd.read_excel(file)
            deleted = []

            for name in df.iloc[:,0]:
                for f in list_files(folder):
                    if f.lower().startswith(str(name).lower()):
                        os.remove(os.path.join(folder,f))
                        deleted.append(f)

            st.dataframe(pd.DataFrame(deleted, columns=["Deleted Files"]))

    # ---------- EXCEL FORMULA UPDATER ----------
    elif tool == "Excel Formula Updater":
        st.header("📊 Excel Formula Updater")
        folder = st.text_input("Folder Path")

        if st.button("🚀 Update Formulas"):
            for f in list_files(folder):
                if f.endswith(".xlsx"):
                    wb = load_workbook(os.path.join(folder,f))
                    ws = wb.active
                    for row in range(2, ws.max_row+1):
                        for col in range(1, ws.max_column+1):
                            cell = ws.cell(row=row, column=col)
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                cell.value = Translator(
                                    cell.value,
                                    origin=cell.coordinate
                                ).translate_formula(cell.coordinate)
                    wb.save(os.path.join(folder,f))
            st.success("✅ Formula update completed")

    # =====================================================
# 6️⃣ COURIER COST UPDATER TOOL (FULLY MIRRORING WORKING SCRIPT)
# =====================================================
elif section == "🛠 Master" and tool == "Courier Cost Updater":

    st.title("🚚 Courier Cost Updater Tool")
    
    testing_folder = st.text_input("📂 Testing Folder Path", r"D:\Sahil\Invoices\Python\Nov-2025\tODAY\07")
    cost_folder = st.text_input("📂 Courier Cost Folder Path", r"D:\Sahil\Invoices\Python\Courier Cost Final File ( From Ashwani )")
    
    source_files = live_file_status(testing_folder)
    start_btn = st.button("🚀 Start Updating Cost Data")

    if start_btn:
        progress = st.progress(0)
        logs = st.empty()
        processed, errors = [], []

        # --- Find cost file ---
        cost_file = next((f for f in os.listdir(cost_folder) if f.endswith(".xlsx") and not f.startswith("~$")), None)
        if not cost_file:
            st.error("❌ No cost file found in folder!")
        else:
            cost_file_path = os.path.join(cost_folder, cost_file)

            # --- Sheet1 ---
            df_cost1 = pd.read_excel(cost_file_path, sheet_name="Sheet1")
            df_cost1.columns.values[0] = 'AWB No'
            df_cost1['AWB No'] = df_cost1['AWB No'].astype(str)
            df_cost1_unique = df_cost1.iloc[:, :13].drop_duplicates(subset='AWB No', keep='first')
            cost_lookup1 = df_cost1_unique.set_index('AWB No').to_dict(orient='index')

            # --- Sheet2 ---
            df_cost2 = pd.read_excel(cost_file_path, sheet_name="Sheet2")
            df_cost2.columns.values[0] = 'AWB No'
            df_cost2['AWB No'] = df_cost2['AWB No'].astype(str)
            df_cost2_unique = df_cost2.iloc[:, :13].drop_duplicates(subset='AWB No', keep='first')
            cost_lookup2 = df_cost2_unique.set_index('AWB No').to_dict(orient='index')

            headers = [
                "Courier charged weight", "Courier Zone", "Freight", "RTO", "RTO Discount", "Reverse",
                "COD", "SDL", "Fuel", "QC", "Others", "Gross"
            ]
            col_start = 53  # BA column

            for idx, testing_file in enumerate(source_files, start=1):
                try:
                    testing_file_path = os.path.join(testing_folder, testing_file)
                    wb = openpyxl.load_workbook(testing_file_path)
                    ws = wb.active
                    max_row = ws.max_row

                    # 🔄 Clear BA-BS (old cost data)
                    for row in range(1, max_row + 1):
                        for col in range(col_start, col_start + len(headers)):
                            ws.cell(row=row, column=col).value = None

                    # 🧹 Blank entire row where column A is empty (from row 3 onward)
                    for row in range(3, max_row + 1):
                        if ws.cell(row=row, column=1).value in [None, ""]:
                            for col in range(1, col_start + len(headers)):
                                ws.cell(row=row, column=col).value = None

                    # 📝 Write headers in row 1 (BA-BM)
                    for i, header in enumerate(headers):
                        ws.cell(row=1, column=col_start + i).value = header

                    # 📝 Write cost data into rows 2+
                    for row in range(2, max_row + 1):
                        awb = ws.cell(row=row, column=4).value  # Column D
                        if awb and str(awb).strip() != "":
                            awb_str = str(awb).strip()
                            cost_row = None
                            source_sheet = None

                            # First check Sheet1, then Sheet2
                            if awb_str in cost_lookup1:
                                cost_row = list(cost_lookup1[awb_str].values())
                                source_sheet = "Sheet1"
                            elif awb_str in cost_lookup2:
                                cost_row = list(cost_lookup2[awb_str].values())
                                source_sheet = "Sheet2"

                            # Write cost data if found
                            if cost_row:
                                for col_offset, val in enumerate(cost_row):
                                    ws.cell(row=row, column=col_start + col_offset).value = val

                    # 💾 Save file
                    wb.save(testing_file_path)
                    processed.append(testing_file)
                    logs.write(f"✅ ({idx}/{len(source_files)}) Updated: {testing_file}")
                except Exception as e:
                    errors.append(f"{testing_file} → {e}")
                    logs.write(f"❌ ({idx}/{len(source_files)}) Error: {testing_file} → {e}")

                progress.progress(idx / len(source_files))

            st.success(f"🎉 Courier Cost Update Completed! Processed: {len(processed)} | Errors: {len(errors)}")
            if processed:
                st.dataframe(pd.DataFrame(processed, columns=["Processed"]))
            if errors:
                st.dataframe(pd.DataFrame(errors, columns=["Errors"]))
